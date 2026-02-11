import math
import datetime as dt
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from PIL import Image

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================
st.set_page_config(page_title="SupplyChain AI Starter Kit", layout="wide")

st.markdown(
    """
<style>
.block-container {padding-top: 1.2rem; padding-bottom: 2rem;}

/* Enterprise header */
.enterprise-bar{
  background: linear-gradient(90deg, #0f172a, #1e3a8a);
  color: white;
  border-radius: 10px;
  padding: 14px 16px;
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap: 12px;
  margin-bottom: 16px;
}
.brand{ display:flex; align-items:center; gap:10px; }
.logo{
  width:34px; height:34px; border-radius:9px;
  background: rgba(255,255,255,0.12);
  display:flex; align-items:center; justify-content:center;
  font-weight:700;
  border:1px solid rgba(255,255,255,0.18);
}
.brand-title{ font-weight:700; letter-spacing:0.4px; font-size:14px; }
.brand-sub{ font-size:12px; opacity:0.85; }
.badges{ display:flex; gap:8px; align-items:center; flex-wrap:wrap; }
.badge{
  padding: 3px 10px; border-radius: 999px; font-size: 12px;
  border: 1px solid rgba(255,255,255,0.20);
  background: rgba(255,255,255,0.10);
}

/* KPI grid */
.kpi-wrap{
  display:grid;
  grid-template-columns: repeat(4, minmax(0, 1fr));
  gap: 10px;
  margin-top: 10px;
}
.kpi{
  background: white;
  border: 1px solid #e6e6e6;
  border-radius: 10px;
  padding: 12px 14px;
}
.kpi-label{ font-size: 12px; color:#6b7280; }
.kpi-value{ font-size: 22px; font-weight: 700; margin-top: 2px; }
.kpi-sub{ font-size: 12px; color:#6b7280; margin-top: 2px; }

@media (max-width: 1100px){
  .kpi-wrap{ grid-template-columns: 1fr; }
}

.img-box{
  background:white;
  border:1px solid #e6e6e6;
  border-radius: 10px;
  padding: 10px;
  display:flex;
  align-items:center;
  justify-content:center;
}

.small-note{ color:#6b7280; font-size:0.88rem; margin-top:6px; }
</style>
""",
    unsafe_allow_html=True,
)

# =========================
# COLONNE
# =========================
REQUIRED_COLS = [
    "articolo",
    "consumo_mensile",
    "lead_time_giorni",
    "stock_attuale",
    "criticita",
    "valore_unitario",
]

OPTIONAL_DEFAULTS = {
    "stagionale": "no",              # si/no
    "indice_rotazione": 8.0,         # numero
    "deviazione_standard": None,     # mensile
    "livello_servizio": "medio",     # alto/medio/basso
}

TEMPLATE_HEADERS = REQUIRED_COLS + list(OPTIONAL_DEFAULTS.keys())


# =========================
# HELPERS
# =========================
def business_days_in_month(year: int, month: int) -> int:
    start = dt.date(year, month, 1)
    end = dt.date(year + 1, 1, 1) if month == 12 else dt.date(year, month + 1, 1)
    cur, count = start, 0
    while cur < end:
        if cur.weekday() < 5:
            count += 1
        cur += dt.timedelta(days=1)
    return count


def load_data(file) -> pd.DataFrame:
    if file.name.lower().endswith(".csv"):
        df = pd.read_csv(file, sep=None, engine="python")
    else:
        df = pd.read_excel(file)
    df.columns = [str(c).strip().lower() for c in df.columns]
    df = df.loc[:, ~df.columns.str.contains("^unnamed", case=False, na=False)]
    return df


def validate_columns(df: pd.DataFrame):
    return [c for c in REQUIRED_COLS if c not in df.columns]


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # aggiungi opzionali se mancano
    for k, v in OPTIONAL_DEFAULTS.items():
        if k not in df.columns:
            df[k] = v

    # criticità
    df["criticita"] = df["criticita"].astype(str).str.lower().str.strip()
    df["criticita"] = df["criticita"].replace({"alto": "alta", "medio": "media", "basso": "bassa"})

    # stagionale
    df["stagionale"] = df["stagionale"].astype(str).str.lower().str.strip()
    df["stagionale"] = df["stagionale"].replace({"sì": "si", "yes": "si", "y": "si", "true": "si", "1": "si"})
    df.loc[~df["stagionale"].isin(["si", "no"]), "stagionale"] = "no"

    # livello servizio
    df["livello_servizio"] = df["livello_servizio"].astype(str).str.lower().str.strip()
    df["livello_servizio"] = df["livello_servizio"].replace({"high": "alto", "medium": "medio", "low": "basso"})
    df.loc[~df["livello_servizio"].isin(["alto", "medio", "basso"]), "livello_servizio"] = "medio"

    # numeri
    for c in ["consumo_mensile", "lead_time_giorni", "stock_attuale", "valore_unitario", "indice_rotazione", "deviazione_standard"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # righe valide
    df = df.dropna(subset=REQUIRED_COLS)
    return df


def z_from_service(level: str) -> float:
    return {"basso": 1.04, "medio": 1.65, "alto": 2.05}.get(level, 1.65)


def crit_factor(crit: str) -> float:
    return {"bassa": 0.9, "media": 1.0, "alta": 1.1}.get(crit, 1.0)


def rotation_factor(rot: float) -> float:
    if pd.isna(rot):
        return 1.0
    if rot >= 12:
        return 1.10
    if rot >= 6:
        return 1.00
    return 0.90


def season_factor(stagionale: str) -> float:
    return 1.15 if stagionale == "si" else 1.00


def compute_metrics(df: pd.DataFrame, workdays: int) -> pd.DataFrame:
    out = df.copy()

    out["consumo_giornaliero"] = out["consumo_mensile"] / float(workdays)
    out["domanda_lt"] = out["consumo_giornaliero"] * out["lead_time_giorni"]

    # deviazione_standard: mensile -> giornaliera su giorni lavorativi
    out["sigma_daily"] = out["deviazione_standard"] / math.sqrt(workdays)
    out["sqrt_lt"] = out["lead_time_giorni"].apply(lambda x: math.sqrt(max(float(x), 0.0)))

    out["z"] = out["livello_servizio"].apply(z_from_service)
    out["ss_base"] = out["z"] * out["sigma_daily"] * out["sqrt_lt"]

    out["fatt_stag"] = out["stagionale"].apply(season_factor)
    out["fatt_rot"] = out["indice_rotazione"].apply(rotation_factor)
    out["fatt_crit"] = out["criticita"].apply(crit_factor)

    def ss_final(row):
        # fallback se deviazione_standard non presente
        if pd.isna(row["deviazione_standard"]) or pd.isna(row["sigma_daily"]):
            base_pct = 0.50 if row["criticita"] == "alta" else 0.30 if row["criticita"] == "media" else 0.15
            return row["domanda_lt"] * base_pct * row["fatt_stag"] * row["fatt_rot"]
        return max(0.0, row["ss_base"] * row["fatt_stag"] * row["fatt_rot"] * row["fatt_crit"])

    out["scorta_sicurezza"] = out.apply(ss_final, axis=1)

    out["punto_riordino"] = (out["domanda_lt"] + out["scorta_sicurezza"]).apply(lambda x: int(math.ceil(x)))
    out["qty_suggerita"] = (out["punto_riordino"] - out["stock_attuale"]).clip(lower=0).apply(lambda x: int(math.ceil(x)))

    def risk(r):
        if r["stock_attuale"] < r["domanda_lt"]:
            return "alto"
        if r["stock_attuale"] < r["punto_riordino"]:
            return "medio"
        return "basso"

    out["rischio_stockout"] = out.apply(risk, axis=1)

    out["valore_unitario"] = out["valore_unitario"].round(2)
    out["valore_ordine_suggerito"] = (out["qty_suggerita"] * out["valore_unitario"]).round(2)
    out["capitale_immobilizzato"] = (out["stock_attuale"] * out["valore_unitario"]).round(2)

    return out


def genera_prompt(row: pd.Series, year: int, month: int, workdays: int) -> str:
    return f"""
Agisci come responsabile supply chain di una PMI.

Mese: {month:02d}/{year} (giorni lavorativi lun–ven: {workdays})

Articolo: {row['articolo']}
Consumo mensile: {int(row['consumo_mensile'])}
Lead time (giorni): {int(row['lead_time_giorni'])}
Stock attuale: {int(row['stock_attuale'])}
Criticità: {row['criticita']}
Valore unitario (€): {row['valore_unitario']:.2f}

Parametri avanzati:
- Stagionale: {row.get('stagionale', 'no')}
- Indice rotazione: {row.get('indice_rotazione', '')}
- Deviazione standard (mensile): {row.get('deviazione_standard', '')}
- Livello servizio: {row.get('livello_servizio', 'medio')}

Risultati:
- Domanda su lead time: {row['domanda_lt']:.2f}
- Scorta sicurezza: {row['scorta_sicurezza']:.2f}
- Punto riordino: {int(row['punto_riordino'])}
- Qty suggerita: {int(row['qty_suggerita'])}
- Valore ordine suggerito (€): {row['valore_ordine_suggerito']:.2f}
- Rischio stockout: {row['rischio_stockout']}

Richiesta:
1) Dimmi se riordinare o no e perché (pratico e sintetico).
2) Conferma quantità o proponi alternativa motivata.
3) Suggerisci 2 azioni immediate per ridurre rischio stockout.
""".strip()


def format_eur(x: float) -> str:
    try:
        return f"€ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "€ 0,00"


# =========================
# EXCEL HELPERS
# =========================
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def _apply_table_header(ws, headers):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _set_column_widths(ws, headers):
    # larghezze “pronte da PMI”
    widths = {
        "articolo": 14,
        "criticita": 10,
        "stagionale": 10,
        "livello_servizio": 14,
    }
    for i, h in enumerate(headers, start=1):
        w = widths.get(h, 16)
        ws.column_dimensions[get_column_letter(i)].width = w


def _format_numeric_columns(ws, headers, start_row=2, end_row=500):
    # formattazione numerica
    int_cols = {
        "consumo_mensile",
        "lead_time_giorni",
        "stock_attuale",
        "qty_suggerita",
        "punto_riordino",
    }
    money_cols = {
        "valore_unitario",
        "valore_ordine_suggerito",
        "capitale_immobilizzato",
    }
    float_cols = {
        "consumo_giornaliero",
        "domanda_lt",
        "scorta_sicurezza",
        "sigma_daily",
        "ss_base",
        "deviazione_standard",
        "indice_rotazione",
    }

    for col_idx, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        if h in int_cols:
            fmt = "0"
        elif h in money_cols:
            fmt = "0.00"
        elif h in float_cols:
            fmt = "0.00"
        else:
            continue

        for r in range(start_row, end_row + 1):
            ws[f"{col_letter}{r}"].number_format = fmt


def _add_validations(ws, headers):
    # dropdown stagionale si/no
    if "stagionale" in headers:
        col = get_column_letter(headers.index("stagionale") + 1)
        dv = DataValidation(type="list", formula1='"si,no"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")

    # dropdown livello_servizio
    if "livello_servizio" in headers:
        col = get_column_letter(headers.index("livello_servizio") + 1)
        dv = DataValidation(type="list", formula1='"basso,medio,alto"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")

    # dropdown criticita (utile in input)
    if "criticita" in headers:
        col = get_column_letter(headers.index("criticita") + 1)
        dv = DataValidation(type="list", formula1='"bassa,media,alta"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}500")


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Input"

    headers = TEMPLATE_HEADERS
    _apply_table_header(ws, headers)
    _set_column_widths(ws, headers)
    _add_validations(ws, headers)

    # riga esempio
    ws.append(["A001", 100, 10, 20, "alta", 10.50, "no", 8, 15, "medio"])
    _format_numeric_columns(ws, headers, start_row=2, end_row=500)

    # NOTE sheet
    ws2 = wb.create_sheet("Note")
    ws2.append(["Campo", "Descrizione"])
    ws2.cell(1, 1).font = HEADER_FONT
    ws2.cell(1, 2).font = HEADER_FONT
    ws2.cell(1, 1).fill = HEADER_FILL
    ws2.cell(1, 2).fill = HEADER_FILL

    notes = [
        ("consumo_mensile", "Quantità mensili (unità)."),
        ("deviazione_standard", "Deviazione standard mensile dei consumi (unità). Se vuota → fallback su % domanda_LT."),
        ("indice_rotazione", "Indice di rotazione (es. 4=slow mover, 12=fast mover)."),
        ("stagionale", "si/no. Se 'si' aumenta scorta di sicurezza."),
        ("livello_servizio", "basso/medio/alto → Z-score per scorta di sicurezza."),
        ("criticita", "bassa/media/alta → micro-fattore correttivo scorta."),
    ]
    for r in notes:
        ws2.append(list(r))
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 90

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_results_xlsx(df_input: pd.DataFrame, df_output: pd.DataFrame) -> bytes:
    wb = Workbook()

    # INPUT
    ws_in = wb.active
    ws_in.title = "Input"

    input_headers = TEMPLATE_HEADERS
    _apply_table_header(ws_in, input_headers)
    _set_column_widths(ws_in, input_headers)
    _add_validations(ws_in, input_headers)

    for _, row in df_input.reindex(columns=input_headers).iterrows():
        ws_in.append(row.tolist())
    _format_numeric_columns(ws_in, input_headers, start_row=2, end_row=max(2, 1 + len(df_input)))

    # OUTPUT
    ws_out = wb.create_sheet("Output")

    output_headers = list(df_output.columns)
    _apply_table_header(ws_out, output_headers)
    _set_column_widths(ws_out, output_headers)

    for _, row in df_output.iterrows():
        ws_out.append(row.tolist())

    _format_numeric_columns(ws_out, output_headers, start_row=2, end_row=max(2, 1 + len(df_output)))

    # leggero auto-width su output
    for i, h in enumerate(output_headers, start=1):
        if h not in ("articolo", "rischio_stockout"):
            ws_out.column_dimensions[get_column_letter(i)].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =========================
# HEADER ENTERPRISE
# =========================
st.markdown(
    """
<div class="enterprise-bar">
  <div class="brand">
    <div class="logo">SC</div>
    <div>
      <div class="brand-title">SUPPLYCHAIN AI DECISION SUPPORT</div>
      <div class="brand-sub">PMI Edition • Prompt-ready • Excel-friendly</div>
    </div>
  </div>
  <div class="badges">
    <div class="badge">Version 1.0</div>
    <div class="badge">Business Days (Mon–Fri)</div>
    <div class="badge">Advanced Safety Stock</div>
  </div>
</div>
""",
    unsafe_allow_html=True,
)

# =========================
# HERO
# =========================
left, right = st.columns([1.35, 1])

with left:
    with st.container(border=True):
        st.markdown("### SupplyChain AI Starter Kit")
        st.markdown(
            "Strumento decisionale per Responsabili Logistica e Supply Chain. "
            "Carica i dati → ottieni priorità di riordino → usa l’AI per decidere meglio."
        )
        st.markdown(
            """
- Punto di riordino e rischio stockout  
- Scorta sicurezza con livello servizio / deviazione standard / stagionalità / rotazione  
- KPI e prompt AI pronti all’uso  
"""
        )
        st.markdown(
            '<div class="small-note">Nota: <b>deviazione_standard</b> è assunta mensile e convertita in giornaliera con √(giorni lavorativi).</div>',
            unsafe_allow_html=True
        )

with right:
    with st.container(border=True):
        img_path = Path("assets/logistic_manager_future.png")
        if img_path.exists():
            img = Image.open(img_path)
            target_height = 220
            w, h = img.size
            ratio = target_height / h
            new_w = int(w * ratio)
            img_resized = img.resize((new_w, target_height))
            st.markdown("<div class='img-box'>", unsafe_allow_html=True)
            st.image(img_resized)
            st.markdown("</div>", unsafe_allow_html=True)
        else:
            st.warning("Immagine non trovata: assets/logistic_manager_future.png")

st.divider()

# =========================
# TEMPLATE DOWNLOAD (sempre visibile)
# =========================
st.download_button(
    "⬇️ Scarica Template Excel (aggiornato)",
    data=build_template_xlsx(),
    file_name="SupplyChain_AI_Template_v1.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.markdown('<div class="small-note">Nel template trovi anche un foglio “Note” con spiegazioni.</div>', unsafe_allow_html=True)

st.divider()

# =========================
# SIDEBAR
# =========================
st.sidebar.header("Impostazioni")
today = dt.date.today()
year = int(st.sidebar.number_input("Anno", 2020, 2100, today.year))
month = int(st.sidebar.selectbox("Mese", list(range(1, 13)), index=today.month - 1))
workdays = business_days_in_month(year, month)
st.sidebar.caption(f"Giorni lavorativi (lun–ven): {workdays}")

uploaded = st.sidebar.file_uploader("Carica CSV o Excel", type=["csv", "xlsx", "xls"])

st.sidebar.caption("Obbligatorie: articolo, consumo_mensile, lead_time_giorni, stock_attuale, criticita, valore_unitario")
st.sidebar.caption("Opzionali: stagionale (si/no), indice_rotazione, deviazione_standard, livello_servizio (alto/medio/basso)")

# =========================
# KPI (placeholder)
# =========================
st.markdown("### KPI sintetici")

if not uploaded:
    st.markdown(
        """
<div class="kpi-wrap">
  <div class="kpi"><div class="kpi-label">Articoli caricati</div><div class="kpi-value">—</div><div class="kpi-sub">Carica un file</div></div>
  <div class="kpi"><div class="kpi-label">Rischio stockout (alto)</div><div class="kpi-value">—</div><div class="kpi-sub">Carica un file</div></div>
  <div class="kpi"><div class="kpi-label">Capitale immobilizzato</div><div class="kpi-value">—</div><div class="kpi-sub">Carica un file</div></div>
  <div class="kpi"><div class="kpi-label">Ordine suggerito (€)</div><div class="kpi-value">—</div><div class="kpi-sub">Carica un file</div></div>
</div>
""",
        unsafe_allow_html=True,
    )
    st.info("Carica un file per visualizzare analisi e KPI.")
    st.stop()

# =========================
# PROCESSING
# =========================
df_raw = load_data(uploaded)
missing = validate_columns(df_raw)
if missing:
    st.error(f"Colonne mancanti: {missing}")
    st.stop()

df = normalize_df(df_raw)
if df.empty:
    st.error("Nessuna riga valida trovata. Controlla intestazioni e valori numerici.")
    st.stop()

metrics = compute_metrics(df, workdays)

# =========================
# KPI CALCOLATI
# =========================
items_count = int(metrics.shape[0])
high_risk_count = int((metrics["rischio_stockout"] == "alto").sum())
capital_locked = float(metrics["capitale_immobilizzato"].sum())
order_suggested_eur = float(metrics["valore_ordine_suggerito"].sum())

st.markdown(
    f"""
<div class="kpi-wrap">
  <div class="kpi">
    <div class="kpi-label">Articoli caricati</div>
    <div class="kpi-value">{items_count}</div>
    <div class="kpi-sub">righe valide nel file</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Rischio stockout (alto)</div>
    <div class="kpi-value">{high_risk_count}</div>
    <div class="kpi-sub">articoli sotto domanda LT</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Capitale immobilizzato</div>
    <div class="kpi-value">{format_eur(capital_locked)}</div>
    <div class="kpi-sub">stock_attuale × valore_unitario</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Ordine suggerito (€)</div>
    <div class="kpi-value">{format_eur(order_suggested_eur)}</div>
    <div class="kpi-sub">qty_suggerita × valore_unitario</div>
  </div>
</div>
""",
    unsafe_allow_html=True,
)

st.divider()

# =========================
# DOWNLOAD OUTPUT EXCEL (Input + Output)
# =========================
st.download_button(
    "⬇️ Scarica Risultati Excel (Input + Output)",
    data=build_results_xlsx(df[TEMPLATE_HEADERS], metrics),
    file_name="SupplyChain_AI_Results_v1.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.divider()

# =========================
# TOP 10 RISCHIO ALTO
# =========================
st.markdown("### Top 10 • Rischio alto")
top10 = metrics[metrics["rischio_stockout"] == "alto"].copy()
if top10.empty:
    st.info("Nessun articolo in rischio ALTO.")
else:
    top10 = top10.sort_values(["valore_ordine_suggerito", "valore_unitario"], ascending=[False, False]).head(10)
    show_cols = [
        "articolo", "consumo_mensile", "lead_time_giorni", "stock_attuale",
        "criticita", "livello_servizio", "stagionale", "indice_rotazione",
        "deviazione_standard", "scorta_sicurezza", "punto_riordino",
        "qty_suggerita", "valore_ordine_suggerito"
    ]
    st.dataframe(top10[show_cols], use_container_width=True)

# =========================
# ANALISI COMPLETA
# =========================
st.markdown("### Analisi completa")
st.dataframe(metrics, use_container_width=True)

# =========================
# PROMPT AI
# =========================
st.markdown("### Prompt AI decisionale")
art = st.selectbox("Seleziona articolo", metrics["articolo"].astype(str).tolist())
row = metrics[metrics["articolo"].astype(str) == art].iloc[0]
st.text_area("Prompt pronto (copia e incolla in ChatGPT)", genera_prompt(row, year, month, workdays), height=280)

